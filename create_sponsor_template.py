import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_sponsor_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sponsor Input"
    
    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    subheader_font = Font(bold=True, size=12, color="000000")
    label_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    
    # Project Metadata Section
    row = 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "PROJECT METADATA"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")
    
    row += 2
    metadata_fields = [
        ("Project Title", "Enter project name"),
        ("Location Country", "Enter country"),
        ("Status", "0. Archived | 1. Site Identified | 1a. Site Reviewed | 2. Test Fit / Power | 3. Deal Development | 4. Under Offer | 5. Exclusivity / Option Secured | 6. Design & Procurement | 6a. End User Engagement | 7. On Site | 8. Live"),
        ("CapEx Currency", "USD/EUR/GBP"),
        ("Cost per kWh Currency", "USD/EUR/GBP"),
        ("Gross IT Load (MW)", "Enter value"),
        ("PUE", "0.95 - 1.5"),
        ("Gross Monthly Rent (per kWh)", "Enter value"),
        ("OPEX as %", "Enter percentage"),
        ("Construction Start Date", "MM/YYYY"),
        ("Construction Duration (Months)", "Enter months"),
    ]
    
    for field, placeholder in metadata_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = placeholder
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # CapEx Section
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "CAPEX BREAKDOWN"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")
    
    row += 2
    ws[f'A{row}'] = "Build (CapEx Internal)"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws[f'B{row}'] = "Per MW Cost"
    ws[f'B{row}'].font = label_font
    ws[f'C{row}'] = "Total Cost"
    ws[f'C{row}'].font = label_font
    ws[f'D{row}'] = "Currency"
    ws[f'D{row}'].font = label_font
    
    row += 1
    ws[f'A{row}'] = "Option 1: Enter Per MW Cost"
    ws[f'B{row}'] = 0
    ws[f'C{row}'] = "=B{0}*B6".format(row)  # References Gross IT Load
    ws[f'D{row}'] = "=B4"  # References CapEx Currency
    
    row += 2
    ws[f'A{row}'] = "Build (CapEx Market)"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws[f'B{row}'] = "Per MW Cost"
    ws[f'B{row}'].font = label_font
    ws[f'C{row}'] = "Total Cost"
    ws[f'C{row}'].font = label_font
    ws[f'D{row}'] = "Currency"
    ws[f'D{row}'].font = label_font
    
    row += 1
    ws[f'A{row}'] = "Option 1: Enter Per MW Cost"
    ws[f'B{row}'] = 0
    ws[f'C{row}'] = "=B{0}*B6".format(row)  # References Gross IT Load
    ws[f'D{row}'] = "=B4"  # References CapEx Currency
    
    row += 2
    ws[f'A{row}'] = "Option 2: Detailed Breakdown"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    capex_items = [
        "Architectural",
        "Civil & Structural",
        "Substations",
        "Underground Utilities",
        "Crusader Modules",
        "BESS",
        "Main Contractor Installation",
        "Photovoltaic System",
        "Landscaping",
        "Preliminaries General",
        "Main Contractor Margin"
    ]
    
    for item in capex_items:
        ws[f'A{row}'] = item
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = 0
        ws[f'C{row}'] = "Currency"
        ws[f'D{row}'] = "=B4"  # References CapEx Currency
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
        row += 1
    
    # Total row
    row += 1
    ws[f'A{row}'] = "TOTAL (Detailed)"
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FF0000")
    ws[f'B{row}'] = f"=SUM(B{row-11}:B{row-1})"
    ws[f'B{row}'].font = Font(bold=True, size=11, color="FF0000")
    
    # Financing Section
    row += 3
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "FINANCING"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")
    
    row += 2
    financing_fields = [
        ("Development Finance (%)", "Enter percentage"),
        ("Construction to Perm (%)", "Enter percentage"),
        ("Development Finance Rate (%)", "Enter rate"),
        ("Perm Debt Rate (%)", "Enter rate"),
        ("ECA Debt Available", "Yes/No"),
        ("ECA Debt Amount", "Enter amount"),
    ]
    
    for field, placeholder in financing_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = placeholder
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # Operational Metrics Section
    row += 2
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "OPERATIONAL METRICS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")
    
    row += 2
    operational_fields = [
        ("Target Yield (%)", "Enter percentage"),
        ("Committed Take (MW)", "Enter MW"),
        ("Option Take (MW)", "Enter MW"),
        ("PPA Structure", "Enter structure details"),
    ]
    
    for field, placeholder in operational_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = placeholder
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    
    # Save the file
    output_path = os.path.join("static", "templates", "Project_Sponsor_Input_Template.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Template created successfully at: {output_path}")

if __name__ == "__main__":
    create_sponsor_template()